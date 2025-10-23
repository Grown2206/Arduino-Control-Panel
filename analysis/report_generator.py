import io
import base64
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from datetime import datetime

# ReportLab imports (PDF)
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# OpenPyXL imports (Excel)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ReportGenerator:
    """
    Professioneller Report-Generator für umfassende Test-Analysen.
    Unterstützt: HTML, PDF und Excel mit erweiterten Visualisierungen.
    """
    
    # Farben für verschiedene Severity-Level
    COLORS = {
        'critical': '#e74c3c',
        'high': '#e67e22',
        'medium': '#f39c12',
        'low': '#3498db',
        'success': '#2ecc71',
        'info': '#3498db',
        'warning': '#f39c12'
    }
    
    @staticmethod
    def generate_html(run_details: dict, analysis: dict) -> str:
        """
        Generiert umfassenden HTML-Bericht mit professioneller Formatierung.
        """
        quality = analysis.get('quality_metrics', {})
        perf_rating = analysis.get('performance_rating', {})
        cycle_stats = analysis.get('cycle_analysis', {})
        anomaly_details = analysis.get('anomaly_details', {})
        stat_summary = analysis.get('statistical_summary', {})
        
        # Header mit Zusammenfassung
        overall_score = quality.get('overall_score', 0)
        rating = perf_rating.get('rating', 'N/A')
        pass_fail = perf_rating.get('pass_fail', 'N/A')
        
        # Farbe basierend auf Score
        if overall_score >= 85:
            score_color = ReportGenerator.COLORS['success']
        elif overall_score >= 70:
            score_color = ReportGenerator.COLORS['warning']
        else:
            score_color = ReportGenerator.COLORS['critical']
        
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #1e1e1e; color: #ecf0f1; margin: 20px; }}
                h1 {{ color: #3498db; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
                h2 {{ color: #e74c3c; margin-top: 30px; border-left: 4px solid #e74c3c; padding-left: 10px; }}
                h3 {{ color: #95a5a6; margin-top: 20px; }}
                
                .header-box {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                              padding: 25px; border-radius: 10px; margin-bottom: 30px; text-align: center; }}
                .score-display {{ font-size: 72px; font-weight: bold; color: {score_color}; margin: 10px 0; }}
                .rating-text {{ font-size: 24px; color: #ecf0f1; }}
                .pass-fail {{ font-size: 28px; font-weight: bold; 
                             color: {'#2ecc71' if pass_fail == 'BESTANDEN' else '#e74c3c'}; 
                             padding: 10px; margin-top: 10px; }}
                
                .info-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); 
                             gap: 15px; margin: 20px 0; }}
                .info-card {{ background: #2c3e50; padding: 15px; border-radius: 8px; 
                             border-left: 4px solid #3498db; }}
                .info-label {{ font-size: 12px; color: #95a5a6; text-transform: uppercase; }}
                .info-value {{ font-size: 20px; font-weight: bold; color: #ecf0f1; margin-top: 5px; }}
                
                .metrics-table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                .metrics-table th {{ background: #34495e; color: #ecf0f1; padding: 12px; text-align: left; }}
                .metrics-table td {{ background: #2c3e50; color: #ecf0f1; padding: 10px; border-bottom: 1px solid #34495e; }}
                
                .anomaly-box {{ background: #2c3e50; padding: 15px; border-radius: 8px; margin: 10px 0;
                               border-left: 4px solid #e74c3c; }}
                .anomaly-critical {{ border-left-color: #e74c3c; }}
                .anomaly-high {{ border-left-color: #e67e22; }}
                .anomaly-medium {{ border-left-color: #f39c12; }}
                
                .recommendation {{ background: #34495e; padding: 12px; margin: 8px 0; 
                                  border-radius: 5px; border-left: 3px solid #3498db; }}
                
                .stat-summary {{ background: #2c3e50; padding: 20px; border-radius: 8px; margin: 20px 0; }}
                
                .progress-bar {{ height: 30px; background: #34495e; border-radius: 15px; 
                                overflow: hidden; margin: 10px 0; }}
                .progress-fill {{ height: 100%; background: linear-gradient(90deg, #3498db, #2ecc71); 
                                 display: flex; align-items: center; justify-content: center; 
                                 color: white; font-weight: bold; transition: width 0.3s; }}
                
                .star-rating {{ font-size: 30px; color: #f39c12; }}
            </style>
        </head>
        <body>
            <div class="header-box">
                <h1 style="color: white; border: none; margin: 0;">Prüfbericht: {run_details.get('name', 'Unbekannt')}</h1>
                <div class="score-display">{overall_score:.1f}</div>
                <div class="rating-text">{rating}</div>
                <div class="star-rating">{'★' * perf_rating.get('star_rating', 0)}{'☆' * (5 - perf_rating.get('star_rating', 0))}</div>
                <div class="pass-fail">{pass_fail}</div>
            </div>
        """
        
        # Basis-Informationen
        html += f"""
            <h2>📋 Testlauf-Informationen</h2>
            <div class="info-grid">
                <div class="info-card">
                    <div class="info-label">Testlauf-ID</div>
                    <div class="info-value">#{run_details.get('id', '-')}</div>
                </div>
                <div class="info-card">
                    <div class="info-label">Sequenz</div>
                    <div class="info-value">{run_details.get('sequence_name', '-')}</div>
                </div>
                <div class="info-card">
                    <div class="info-label">Dauer</div>
                    <div class="info-value">{run_details.get('duration', 0):.2f} s</div>
                </div>
                <div class="info-card">
                    <div class="info-label">Zyklen</div>
                    <div class="info-value">{run_details.get('cycles', 0)}</div>
                </div>
                <div class="info-card">
                    <div class="info-label">Status</div>
                    <div class="info-value">{run_details.get('status', '-')}</div>
                </div>
                <div class="info-card">
                    <div class="info-label">Startzeit</div>
                    <div class="info-value" style="font-size: 14px;">{run_details.get('start_time', '-')}</div>
                </div>
            </div>
        """
        
        # Qualitätsmetriken mit Progress Bars
        html += f"""
            <h2>📊 Qualitätsmetriken</h2>
            <div style="margin: 20px 0;">
                <div style="margin-bottom: 15px;">
                    <div style="display: flex; justify-content: space-between; margin-bottom: 5px;">
                        <span>Konsistenz</span>
                        <span>{quality.get('consistency_score', 0):.1f}%</span>
                    </div>
                    <div class="progress-bar">
                        <div class="progress-fill" style="width: {quality.get('consistency_score', 0)}%;">
                            {quality.get('consistency_score', 0):.1f}%
                        </div>
                    </div>
                </div>
                
                <div style="margin-bottom: 15px;">
                    <div style="display: flex; justify-content: space-between; margin-bottom: 5px;">
                        <span>Zuverlässigkeit</span>
                        <span>{quality.get('reliability_score', 0):.1f}%</span>
                    </div>
                    <div class="progress-bar">
                        <div class="progress-fill" style="width: {quality.get('reliability_score', 0)}%;">
                            {quality.get('reliability_score', 0):.1f}%
                        </div>
                    </div>
                </div>
                
                <div style="margin-bottom: 15px;">
                    <div style="display: flex; justify-content: space-between; margin-bottom: 5px;">
                        <span>Präzision (Jitter)</span>
                        <span>{quality.get('jitter_score', 0):.1f}%</span>
                    </div>
                    <div class="progress-bar">
                        <div class="progress-fill" style="width: {quality.get('jitter_score', 0)}%;">
                            {quality.get('jitter_score', 0):.1f}%
                        </div>
                    </div>
                </div>
            </div>
        """
        
        # Zykluszeit-Analyse
        html += f"""
            <h2>⏱️ Zykluszeit-Analyse</h2>
            <table class="metrics-table">
                <tr>
                    <th>Metrik</th>
                    <th>Wert</th>
                    <th>Interpretation</th>
                </tr>
                <tr>
                    <td>Durchschnitt</td>
                    <td>{cycle_stats.get('avg', 0):.2f} ms</td>
                    <td>Mittlere Ausführungszeit pro Zyklus</td>
                </tr>
                <tr>
                    <td>Median</td>
                    <td>{cycle_stats.get('median', 0):.2f} ms</td>
                    <td>Typische Zykluszeit (50. Perzentil)</td>
                </tr>
                <tr>
                    <td>Standardabweichung</td>
                    <td>{cycle_stats.get('std', 0):.2f} ms</td>
                    <td>Schwankungsbereich</td>
                </tr>
                <tr>
                    <td>Min / Max</td>
                    <td>{cycle_stats.get('min', 0):.2f} / {cycle_stats.get('max', 0):.2f} ms</td>
                    <td>Schnellster / Langsamster Zyklus</td>
                </tr>
                <tr>
                    <td>IQR (Interquartilsabstand)</td>
                    <td>{cycle_stats.get('iqr', 0):.2f} ms</td>
                    <td>Mittlere 50% Streuung</td>
                </tr>
                <tr>
                    <td>Variationskoeffizient</td>
                    <td>{cycle_stats.get('cv', 0):.2f}%</td>
                    <td>Relative Variabilität</td>
                </tr>
                <tr>
                    <td>Stabilität</td>
                    <td style="color: {ReportGenerator._get_stability_color(cycle_stats.get('stability', 0))}; font-weight: bold;">
                        {cycle_stats.get('stability', 0):.1f}%
                    </td>
                    <td>{ReportGenerator._get_stability_interpretation(cycle_stats.get('stability', 0))}</td>
                </tr>
                <tr>
                    <td>Trend</td>
                    <td style="color: {ReportGenerator._get_trend_color(cycle_stats.get('trend', 'stabil'))}; font-weight: bold;">
                        {cycle_stats.get('trend', 'N/A').upper()}
                    </td>
                    <td>R² = {cycle_stats.get('r_squared', 0):.4f}</td>
                </tr>
            </table>
        """
        
        # Anomalie-Analyse
        total_anomalies = anomaly_details.get('total_anomalies', 0)
        if total_anomalies > 0:
            html += f"""
                <h2>⚠️ Anomalie-Analyse</h2>
                <div class="info-grid">
                    <div class="info-card" style="border-left-color: #e74c3c;">
                        <div class="info-label">Gesamt-Anomalien</div>
                        <div class="info-value">{total_anomalies}</div>
                    </div>
                    <div class="info-card" style="border-left-color: #e74c3c;">
                        <div class="info-label">Zyklus-Anomalien</div>
                        <div class="info-value">{anomaly_details.get('cycle_anomalies', 0)}</div>
                    </div>
                    <div class="info-card" style="border-left-color: #f39c12;">
                        <div class="info-label">Schritt-Anomalien</div>
                        <div class="info-value">{anomaly_details.get('step_anomalies', 0)}</div>
                    </div>
                    <div class="info-card" style="border-left-color: #e67e22;">
                        <div class="info-label">Anomalierate</div>
                        <div class="info-value">{anomaly_details.get('anomaly_rate', 0):.1f}%</div>
                    </div>
                </div>
            """
            
            # Kritischste Anomalie
            # (Hinweis: 'most_severe_anomaly' wird vom aktuellen TrendAnalyzer evtl. nicht geliefert,
            # aber wir lassen den Code defensiv hier)
            most_severe = anomaly_details.get('most_severe_anomaly')
            if most_severe:
                severity_color = ReportGenerator.COLORS.get(most_severe.get('severity', 'medium'), '#f39c12')
                html += f"""
                    <div class="anomaly-box anomaly-{most_severe.get('severity', 'medium')}">
                        <h3 style="margin-top: 0; color: {severity_color};">
                            🔴 Kritischste Anomalie
                        </h3>
                        <p><strong>Zyklus:</strong> #{most_severe.get('cycle', 0)}</p>
                        <p><strong>Zeit:</strong> {most_severe.get('time', 0):.2f} ms</p>
                        <p><strong>Abweichung:</strong> {most_severe.get('deviation_percent', 0):.1f}%</p>
                        <p><strong>Z-Score:</strong> {most_severe.get('z_score', 0):.2f}</p>
                        <p><strong>Schweregrad:</strong> {most_severe.get('severity', 'unbekannt').upper()}</p>
                    </div>
                """
            
            # Problem-Schritte
            problem_steps = anomaly_details.get('problem_steps', [])
            if problem_steps:
                html += "<h3>🎯 Problematische Schritte</h3>"
                for item in problem_steps[:5]:  # Top 5
                    step = item.get('step', 'Unbekannt')
                    data = item.get('data', {})
                    html += f"""
                        <div class="anomaly-box anomaly-medium">
                            <strong>{step}</strong><br>
                            Anomalien: {data.get('count', 0)} ({data.get('percentage', 0):.1f}%)<br>
                            Ø Abweichung: {data.get('avg_deviation', 0):.1f}%
                        </div>
                    """
        
        # Statistische Zusammenfassung
        if stat_summary:
            html += f"""
                <h2>📈 Statistische Zusammenfassung</h2>
                <div class="stat-summary">
                    <div class="info-grid">
                        <div>
                            <div class="info-label">Stichprobengröße</div>
                            <div class="info-value">{stat_summary.get('sample_size', 0)}</div>
                        </div>
                        <div>
                            <div class="info-label">Schiefe (Skewness)</div>
                            <div class="info-value">{stat_summary.get('skewness', 0):.3f}</div>
                        </div>
                        <div>
                            <div class="info-label">Wölbung (Kurtosis)</div>
                            <div class="info-value">{stat_summary.get('kurtosis', 0):.3f}</div>
                        </div>
                        <div>
                            <div class="info-label">Varianz</div>
                            <div class="info-value">{stat_summary.get('variance', 0):.2f}</div>
                        </div>
                    </div>
                    
                    <h4 style="color: #95a5a6; margin-top: 20px;">Perzentile</h4>
                    <table class="metrics-table" style="font-size: 12px;">
                        <tr>
                            <th>P10</th><th>P25</th><th>P50</th><th>P75</th><th>P90</th><th>P95</th><th>P99</th>
                        </tr>
            """
            
            percentiles = stat_summary.get('percentiles', {})
            for p in [10, 25, 50, 75, 90, 95, 99]:
                html += f"<td>{percentiles.get(p, 0):.2f} ms</td>"
            
            html += """
                        </tr>
                    </table>
                </div>
            """
        
        # Empfehlungen
        recommendations = quality.get('recommendations', [])
        if recommendations:
            html += "<h2>💡 Empfehlungen</h2>"
            for rec in recommendations:
                html += f'<div class="recommendation">{rec}</div>'
        
        # Performance-Rating Details
        html += f"""
            <h2>🏆 Performance-Rating</h2>
            <div class="info-grid">
        """
        
        categories = perf_rating.get('categories', {})
        if categories:
            for category, score in categories.items():
                color = ReportGenerator._get_score_color(score)
                html += f"""
                    <div class="info-card" style="border-left-color: {color};">
                        <div class="info-label">{category}</div>
                        <div class="info-value" style="color: {color};">{score:.1f}%</div>
                    </div>
                """
        
        html += "</div>"
        
        # Stärken und Schwächen
        strengths = perf_rating.get('strengths', [])
        weaknesses = perf_rating.get('weaknesses', [])
        
        if strengths:
            html += "<h3 style='color: #2ecc71;'>✅ Stärken</h3><ul>"
            for strength in strengths:
                html += f"<li>{strength}</li>"
            html += "</ul>"
        
        if weaknesses:
            html += "<h3 style='color: #e74c3c;'>⚠️ Verbesserungspotenzial</h3><ul>"
            for weakness in weaknesses:
                html += f"<li>{weakness}</li>"
            html += "</ul>"
        
        # Sensor-Daten
        log_data = run_details.get('log', {})
        sensor_stats = log_data.get('sensors', {})
        
        if sensor_stats:
            html += """
                <h2>🌡️ Sensor-Auswertung</h2>
                <table class="metrics-table">
                    <tr>
                        <th>Sensor</th>
                        <th>Minimum</th>
                        <th>Maximum</th>
                        <th>Durchschnitt</th>
                    </tr>
            """
            
            if 'temp' in sensor_stats:
                s = sensor_stats['temp']
                html += f"""
                    <tr>
                        <td>🌡️ Temperatur</td>
                        <td>{s.get('min', 0):.1f}°C</td>
                        <td>{s.get('max', 0):.1f}°C</td>
                        <td>{s.get('avg', 0):.1f}°C</td>
                    </tr>
                """
            
            if 'humid' in sensor_stats:
                s = sensor_stats['humid']
                html += f"""
                    <tr>
                        <td>💧 Luftfeuchtigkeit</td>
                        <td>{s.get('min', 0):.1f}%</td>
                        <td>{s.get('max', 0):.1f}%</td>
                        <td>{s.get('avg', 0):.1f}%</td>
                    </tr>
                """
            
            html += "</table>"
        
        # Footer
        html += f"""
            <div style="margin-top: 50px; padding-top: 20px; border-top: 2px solid #34495e; 
                        text-align: center; color: #95a5a6; font-size: 12px;">
                <p>Bericht generiert am {datetime.now().strftime('%d.%m.%Y um %H:%M:%S')}</p>
                <p>Drexler Dynamics - Arduino Control Panel</p>
            </div>
        </body>
        </html>
        """
        
        return html
    
    @staticmethod
    def _get_stability_color(stability: float) -> str:
        """Gibt Farbe basierend auf Stabilitätswert zurück."""
        if stability >= 95:
            return '#2ecc71'
        elif stability >= 85:
            return '#3498db'
        elif stability >= 70:
            return '#f39c12'
        else:
            return '#e74c3c'
    
    @staticmethod
    def _get_stability_interpretation(stability: float) -> str:
        """Gibt Interpretation des Stabilitätswerts zurück."""
        if stability >= 95:
            return 'Exzellent - Sehr geringe Schwankungen'
        elif stability >= 85:
            return 'Sehr gut - Akzeptable Schwankungen'
        elif stability >= 70:
            return 'Gut - Moderate Schwankungen'
        elif stability >= 60:
            return 'Befriedigend - Erhöhte Schwankungen'
        else:
            return 'Mangelhaft - Starke Schwankungen'
    
    @staticmethod
    def _get_trend_color(trend: str) -> str:
        """Gibt Farbe basierend auf Trend zurück."""
        if trend == 'STABIL':
            return '#2ecc71'
        elif trend == 'STEIGEND':
            return '#e74c3c'
        else:  # FALLEND
            return '#3498db'
    
    @staticmethod
    def _get_score_color(score: float) -> str:
        """Gibt Farbe basierend auf Score zurück."""
        if score >= 85:
            return '#2ecc71'
        elif score >= 70:
            return '#3498db'
        elif score >= 60:
            return '#f39c12'
        else:
            return '#e74c3c'
    
    @staticmethod
    def _create_charts(analysis: dict, sensors_raw: list = None) -> list:
        """
        Erstellt erweiterte Visualisierungen für den Bericht.
        """
        chart_buffers = []
        
        plt.style.use('dark_background') # Für HTML-Vorschau
        
        # Chart 1: Zykluszeiten mit Anomalien
        chart_buffers.append(ReportGenerator._create_cycle_time_chart(analysis))
        
        # Chart 2: Box-Plot
        chart_buffers.append(ReportGenerator._create_box_plot(analysis))
        
        # Chart 3: Schritt-Analyse
        chart_buffers.append(ReportGenerator._create_step_analysis_chart(analysis))
        
        # Chart 4: Histogramm
        chart_buffers.append(ReportGenerator._create_histogram(analysis))
        
        # Chart 5: Sensor-Daten (falls vorhanden)
        if sensors_raw:
            sensor_chart = ReportGenerator._create_sensor_chart(sensors_raw)
            if sensor_chart:
                chart_buffers.append(sensor_chart)
        
        return chart_buffers
    
    @staticmethod
    def _create_cycle_time_chart(analysis: dict) -> io.BytesIO:
        """Erstellt Liniendiagramm der Zykluszeiten mit markierten Anomalien."""
        cycle_times = analysis.get('raw_cycle_times', [])
        cycle_analysis = analysis.get('cycle_analysis', {})
        anomalies = cycle_analysis.get('anomalies', [])
        
        if not cycle_times:
            return io.BytesIO()
        
        fig, ax = plt.subplots(figsize=(12, 6))
        
        x = np.arange(1, len(cycle_times) + 1)
        
        # Hauptlinie
        ax.plot(x, cycle_times, 'o-', color='#3498db', linewidth=2, markersize=4, label='Zykluszeiten')
        
        # Durchschnitt
        avg = cycle_analysis.get('avg', 0)
        ax.axhline(y=avg, color='#2ecc71', linestyle='--', linewidth=1.5, label=f'Durchschnitt ({avg:.2f} ms)')
        
        # Standardabweichungs-Band
        std = cycle_analysis.get('std', 0)
        ax.fill_between(x, avg - std, avg + std, alpha=0.2, color='#2ecc71', label='±1 Standardabweichung')
        
        # Anomalien markieren
        if anomalies:
            anomaly_cycles = [a['cycle'] for a in anomalies]
            anomaly_times = [a['time'] for a in anomalies]
            
            anomaly_colors = [ReportGenerator.COLORS.get(a.get('severity', 'medium'), '#f39c12') for a in anomalies]
            
            ax.scatter(anomaly_cycles, anomaly_times, s=150, c=anomaly_colors, 
                      marker='X', edgecolors='white', linewidths=1.5, 
                      label='Anomalien', zorder=5)
        
        # Trend-Linie
        if len(cycle_times) > 2:
            z = np.polyfit(x, cycle_times, 1)
            p = np.poly1d(z)
            ax.plot(x, p(x), "--", color='#e74c3c', linewidth=1.5, alpha=0.7, label='Trend')
        
        ax.set_xlabel('Zyklus-Nummer', fontsize=12)
        ax.set_ylabel('Zeit (ms)', fontsize=12)
        ax.set_title('Zykluszeiten-Verlauf mit Anomalieerkennung', fontsize=14, fontweight='bold')
        ax.legend(loc='best', fontsize=10)
        ax.grid(True, alpha=0.3, linestyle='--')
        
        plt.tight_layout()
        
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=120, transparent=True)
        buf.seek(0)
        plt.close(fig)
        
        return buf
    
    @staticmethod
    def _create_box_plot(analysis: dict) -> io.BytesIO:
        """Erstellt Box-Plot zur Verteilungsanalyse."""
        cycle_times = analysis.get('raw_cycle_times', [])
        
        if not cycle_times:
            return io.BytesIO()
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        bp = ax.boxplot([cycle_times], vert=True, patch_artist=True, 
                        labels=['Zykluszeiten'],
                        boxprops=dict(facecolor='#3498db', alpha=0.7),
                        whiskerprops=dict(color='#3498db', linewidth=1.5),
                        capprops=dict(color='#3498db', linewidth=1.5),
                        medianprops=dict(color='#e74c3c', linewidth=2),
                        flierprops=dict(marker='o', markerfacecolor='#e74c3c', 
                                      markersize=8, linestyle='none'))
        
        # Statistiken anzeigen
        cycle_analysis = analysis.get('cycle_analysis', {})
        stat_summary = analysis.get('statistical_summary', {})
        percentiles = stat_summary.get('percentiles', {})

        stats_text = f"""Median: {cycle_analysis.get('median', 0):.2f} ms
Q1: {percentiles.get(25, 0):.2f} ms
Q3: {percentiles.get(75, 0):.2f} ms
IQR: {cycle_analysis.get('iqr', 0):.2f} ms"""
        
        ax.text(1.15, np.median(cycle_times), stats_text, 
               verticalalignment='center', fontsize=10,
               bbox=dict(boxstyle='round', facecolor='#2c3e50', alpha=0.8))
        
        ax.set_ylabel('Zeit (ms)', fontsize=12)
        ax.set_title('Verteilungsanalyse der Zykluszeiten (Box-Plot)', fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3, linestyle='--', axis='y')
        
        plt.tight_layout()
        
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=120, transparent=True)
        buf.seek(0)
        plt.close(fig)
        
        return buf
    
    @staticmethod
    def _create_step_analysis_chart(analysis: dict) -> io.BytesIO:
        """Erstellt Balkendiagramm der Schritt-Zeiten mit Fehlerbalken."""
        step_analysis = analysis.get('step_analysis', {})
        
        if not step_analysis:
            return io.BytesIO()
        
        # Sortiere nach durchschnittlicher Zeit
        sorted_steps = sorted(step_analysis.items(), key=lambda x: x[1].get('avg', 0), reverse=True)
        
        # Nehme Top 10 Schritte
        top_steps = sorted_steps[:10]
        
        labels = [step[:30] + '...' if len(step) > 30 else step for step, _ in top_steps]
        avgs = [data.get('avg', 0) for _, data in top_steps]
        stds = [data.get('std', 0) for _, data in top_steps]
        
        fig, ax = plt.subplots(figsize=(12, 8))
        
        y_pos = np.arange(len(labels))
        # Defensiv prüfen, ob 'anomaly_count' existiert
        colors_list = ['#3498db' if data.get('anomaly_count', 0) == 0 else '#e74c3c' 
                      for _, data in top_steps]
        
        bars = ax.barh(y_pos, avgs, xerr=stds, align='center', 
                      color=colors_list, alpha=0.8, capsize=5)
        
        ax.set_yticks(y_pos)
        ax.set_yticklabels(labels, fontsize=9)
        ax.invert_yaxis()
        ax.set_xlabel('Zeit (ms)', fontsize=12)
        ax.set_title('Top 10 Schritte nach durchschnittlicher Ausführungszeit', 
                    fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3, linestyle='--', axis='x')
        
        # Legende
        from matplotlib.patches import Patch
        legend_elements = [
            Patch(facecolor='#3498db', alpha=0.8, label='Ohne Anomalien'),
            Patch(facecolor='#e74c3c', alpha=0.8, label='Mit Anomalien')
        ]
        ax.legend(handles=legend_elements, loc='best')
        
        plt.tight_layout()
        
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=120, transparent=True)
        buf.seek(0)
        plt.close(fig)
        
        return buf
    
    @staticmethod
    def _create_histogram(analysis: dict) -> io.BytesIO:
        """Erstellt Histogramm der Zykluszeiten-Verteilung."""
        cycle_times = analysis.get('raw_cycle_times', [])
        
        if not cycle_times:
            return io.BytesIO()
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Histogramm
        n, bins, patches = ax.hist(cycle_times, bins=30, color='#3498db', 
                                   alpha=0.7, edgecolor='black')
        
        # Normalverteilungs-Kurve
        cycle_analysis = analysis.get('cycle_analysis', {})
        avg = cycle_analysis.get('avg', 0)
        std = cycle_analysis.get('std', 0)
        
        if std > 0:
            x = np.linspace(min(cycle_times), max(cycle_times), 100)
            try:
                from scipy.stats import norm
                y = norm.pdf(x, avg, std) * len(cycle_times) * (bins[1] - bins[0])
                ax.plot(x, y, 'r--', linewidth=2, label='Normalverteilung')
            except ImportError:
                print("Hinweis: scipy.stats nicht gefunden, Normalverteilung wird nicht gezeichnet.")

        
        ax.axvline(avg, color='#2ecc71', linestyle='--', linewidth=2, label=f'Mittelwert ({avg:.2f} ms)')
        ax.set_xlabel('Zeit (ms)', fontsize=12)
        ax.set_ylabel('Häufigkeit', fontsize=12)
        ax.set_title('Häufigkeitsverteilung der Zykluszeiten', fontsize=14, fontweight='bold')
        ax.legend()
        ax.grid(True, alpha=0.3, linestyle='--', axis='y')
        
        plt.tight_layout()
        
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=120, transparent=True)
        buf.seek(0)
        plt.close(fig)
        
        return buf
    
    @staticmethod
    def _create_sensor_chart(sensors_raw: list) -> io.BytesIO:
        """Erstellt Sensor-Verlaufsdiagramm."""
        if not sensors_raw:
            return None
        
        temps = [(s['time_ms'], s['value']) for s in sensors_raw if s.get('sensor') == 'B24_TEMP' and s.get('value') is not None and s.get('time_ms') is not None]
        humids = [(s['time_ms'], s['value']) for s in sensors_raw if s.get('sensor') == 'B24_HUMIDITY' and s.get('value') is not None and s.get('time_ms') is not None]
        
        if not temps and not humids:
            return None
        
        fig, axes = plt.subplots(2, 1, figsize=(12, 8), sharex=True)
        
        if temps:
            temps.sort(key=lambda x: x[0]) # Sortieren nach Zeit
            times, values = zip(*temps)
            axes[0].plot(times, values, color='#e74c3c', linewidth=2, marker='o', markersize=3)
            axes[0].set_ylabel('Temperatur (°C)', fontsize=12)
            axes[0].set_title('Temperatur-Verlauf', fontsize=14, fontweight='bold')
            axes[0].grid(True, alpha=0.3, linestyle='--')
            axes[0].fill_between(times, values, alpha=0.3, color='#e74c3c')
        
        if humids:
            humids.sort(key=lambda x: x[0]) # Sortieren nach Zeit
            times, values = zip(*humids)
            axes[1].plot(times, values, color='#3498db', linewidth=2, marker='o', markersize=3)
            axes[1].set_ylabel('Luftfeuchtigkeit (%)', fontsize=12)
            axes[1].set_title('Luftfeuchtigkeits-Verlauf', fontsize=14, fontweight='bold')
            axes[1].grid(True, alpha=0.3, linestyle='--')
            axes[1].fill_between(times, values, alpha=0.3, color='#3498db')
        
        axes[1].set_xlabel('Zeit (ms)', fontsize=12)
        
        plt.tight_layout()
        
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=120, transparent=True)
        buf.seek(0)
        plt.close(fig)
        
        return buf
    
    # ===== PDF EXPORT =====
    @staticmethod
    def generate_pdf(run_details: dict, file_path: str):
        """
        Generiert professionellen PDF-Bericht.
        """
        if not REPORTLAB_AVAILABLE:
            raise Exception("ReportLab ist nicht installiert. Bitte installieren Sie: pip install reportlab")
        
        from analysis.trend_analyzer import TrendAnalyzer
        
        log_data = run_details.get('log', {})
        event_log = log_data.get('events', [])
        analysis = TrendAnalyzer.analyze_timing(event_log)
        
        doc = SimpleDocTemplate(
            file_path, 
            pagesize=A4,
            topMargin=2*cm, 
            bottomMargin=2*cm,
            leftMargin=2*cm, 
            rightMargin=2*cm
        )
        
        styles = getSampleStyleSheet()
        
        # Custom Styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#3498db'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        story = []
        
        # Titel
        story.append(Paragraph(f"Prüfbericht: {run_details.get('name', 'Unbekannt')}", title_style))
        story.append(Spacer(1, 0.5*cm))
        
        # Score-Box
        quality = analysis.get('quality_metrics', {})
        perf_rating = analysis.get('performance_rating', {})
        
        score_data = [[
            Paragraph(f"<b>Gesamt-Score</b>", styles['Normal']),
            Paragraph(f"<b><font size=20>{quality.get('overall_score', 0):.1f}</font></b>", styles['Normal']),
            Paragraph(f"<b>{perf_rating.get('rating', 'N/A')}</b>", styles['Normal']),
            Paragraph(f"<b>{perf_rating.get('pass_fail', 'N/A')}</b>", styles['Normal'])
        ]]
        
        score_table = Table(score_data, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
        score_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2c3e50'))
        ]))
        story.append(score_table)
        story.append(Spacer(1, 1*cm))
        
        # Basis-Informationen
        story.append(Paragraph("<b>Testlauf-Informationen</b>", styles['Heading2']))
        
        info_data = [
            ["Testlauf-ID:", str(run_details.get('id', '-'))],
            ["Sequenz:", str(run_details.get('sequence_name', '-'))],
            ["Startzeit:", str(run_details.get('start_time', '-'))],
            ["Endzeit:", str(run_details.get('end_time', '-'))],
            ["Dauer:", f"{run_details.get('duration', 0):.2f} s"],
            ["Zyklen:", str(run_details.get('cycles', 0))],
            ["Status:", run_details.get('status', '-')]
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 12*cm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.black), # PDF Hintergrund ist weiß
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        story.append(info_table)
        story.append(Spacer(1, 1*cm))
        
        # Qualitätsmetriken
        story.append(Paragraph("<b>Qualitätsmetriken</b>", styles['Heading2']))
        
        quality_data = [
            ["Konsistenz:", f"{quality.get('consistency_score', 0):.1f}%"],
            ["Zuverlässigkeit:", f"{quality.get('reliability_score', 0):.1f}%"],
            ["Präzision (Jitter):", f"{quality.get('jitter_score', 0):.1f}%"],
            ["Anomalierate:", f"{quality.get('anomaly_rate', 0):.1f}%"]
        ]
        
        quality_table = Table(quality_data, colWidths=[6*cm, 10*cm])
        quality_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        story.append(quality_table)
        story.append(Spacer(1, 1*cm))
        
        # Empfehlungen
        recommendations = quality.get('recommendations', [])
        if recommendations:
            story.append(Paragraph("<b>Empfehlungen</b>", styles['Heading2']))
            for rec in recommendations:
                story.append(Paragraph(f"• {rec}", styles['Normal']))
                story.append(Spacer(1, 0.2*cm))
        
        story.append(PageBreak())
        
        # Charts
        story.append(Paragraph("<b>Visualisierungen</b>", styles['Heading2']))
        charts = ReportGenerator._create_charts(analysis, log_data.get('sensors_raw'))
        
        for chart_buf in charts:
            story.append(Image(chart_buf, width=16*cm, height=10*cm))
            story.append(Spacer(1, 0.5*cm))
        
        # Build PDF
        doc.build(story)
    
    # ===== EXCEL EXPORT =====
    @staticmethod
    def generate_excel(run_details: dict, file_path: str):
        """
        Generiert umfassenden Excel-Bericht mit mehreren Arbeitsblättern.
        """
        if not EXCEL_AVAILABLE:
            raise Exception("openpyxl ist nicht installiert. Bitte installieren Sie: pip install openpyxl")
        
        from analysis.trend_analyzer import TrendAnalyzer
        
        log_data = run_details.get('log', {})
        event_log = log_data.get('events', [])
        analysis = TrendAnalyzer.analyze_timing(event_log)
        
        wb = Workbook()
        
        # === Sheet 1: Zusammenfassung ===
        ws_summary = wb.active
        ws_summary.title = "Zusammenfassung"
        
        # Header
        ws_summary['A1'] = f"Prüfbericht: {run_details.get('name', '')}"
        ws_summary['A1'].font = Font(size=16, bold=True, color="3498DB")
        ws_summary.merge_cells('A1:E1')
        
        # Score-Bereich
        quality = analysis.get('quality_metrics', {})
        perf_rating = analysis.get('performance_rating', {})
        
        ws_summary['A3'] = "Gesamt-Score"
        ws_summary['B3'] = quality.get('overall_score', 0)
        ws_summary['C3'] = perf_rating.get('rating', 'N/A')
        ws_summary['D3'] = perf_rating.get('pass_fail', 'N/A')
        
        # Styling für Score
        for col in ['A', 'B', 'C', 'D']:
            ws_summary[f'{col}3'].font = Font(size=14, bold=True)
            ws_summary[f'{col}3'].alignment = Alignment(horizontal='center')
            ws_summary[f'{col}3'].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            ws_summary[f'{col}3'].font = Font(size=14, bold=True, color="FFFFFF")
        
        # Basis-Informationen
        current_row = 5
        ws_summary[f'A{current_row}'] = "Testlauf-Informationen"
        ws_summary[f'A{current_row}'].font = Font(size=12, bold=True)
        current_row += 1
        
        info_data = [
            ("Testlauf-ID", run_details.get('id')),
            ("Sequenz", run_details.get('sequence_name')),
            ("Startzeit", run_details.get('start_time')),
            ("Endzeit", run_details.get('end_time')),
            ("Dauer (s)", f"{run_details.get('duration', 0):.2f}"),
            ("Zyklen", run_details.get('cycles', 0)),
            ("Status", run_details.get('status'))
        ]
        
        for key, value in info_data:
            ws_summary[f'A{current_row}'] = key
            ws_summary[f'B{current_row}'] = str(value) if value is not None else ""
            ws_summary[f'A{current_row}'].font = Font(bold=True)
            current_row += 1
        
        # === Sheet 2: Detaillierte Analyse ===
        ws_analysis = wb.create_sheet("Detaillierte Analyse")
        
        ws_analysis['A1'] = "Zykluszeit-Analyse"
        ws_analysis['A1'].font = Font(size=14, bold=True)
        
        cycle_stats = analysis.get('cycle_analysis', {})
        
        analysis_data = [
            ("Durchschnitt (ms)", f"{cycle_stats.get('avg', 0):.2f}"),
            ("Median (ms)", f"{cycle_stats.get('median', 0):.2f}"),
            ("Standardabweichung (ms)", f"{cycle_stats.get('std', 0):.2f}"),
            ("Min (ms)", f"{cycle_stats.get('min', 0):.2f}"),
            ("Max (ms)", f"{cycle_stats.get('max', 0):.2f}"),
            ("IQR (ms)", f"{cycle_stats.get('iqr', 0):.2f}"),
            ("Variationskoeffizient (%)", f"{cycle_stats.get('cv', 0):.2f}"),
            ("Stabilität (%)", f"{cycle_stats.get('stability', 0):.1f}"),
            ("Trend", cycle_stats.get('trend', 'N/A')),
            ("Trend R²", f"{cycle_stats.get('r_squared', 0):.4f}"),
            ("Anomalien", len(cycle_stats.get('anomalies', [])))
        ]
        
        row = 3
        for key, value in analysis_data:
            ws_analysis[f'A{row}'] = key
            ws_analysis[f'B{row}'] = value
            ws_analysis[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # === Sheet 3: Anomalien ===
        ws_anomalies = wb.create_sheet("Anomalien")
        
        ws_anomalies['A1'] = "Anomalie-Bericht"
        ws_anomalies['A1'].font = Font(size=14, bold=True)
        
        anomalies = cycle_stats.get('anomalies', [])
        
        if anomalies:
            headers = ["Zyklus", "Zeit (ms)", "Abweichung (%)", "Z-Score", "Schweregrad"]
            for col, header in enumerate(headers, start=1):
                cell = ws_anomalies.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            row = 4
            for anomaly in anomalies:
                ws_anomalies[f'A{row}'] = anomaly.get('cycle', 0)
                ws_anomalies[f'B{row}'] = f"{anomaly.get('time', 0):.2f}"
                ws_anomalies[f'C{row}'] = f"{anomaly.get('deviation', 0):.1f}" #deviation statt deviation_percent
                ws_anomalies[f'D{row}'] = f"{anomaly.get('z_score', 0):.2f}" # z_score wird evtl. nicht geliefert
                ws_anomalies[f'E{row}'] = anomaly.get('severity', 'unbekannt').upper()
                
                # Farbe basierend auf Schweregrad (defensiv)
                severity = anomaly.get('severity', 'medium')
                if severity == 'critical' or severity == 'kritisch':
                    color = "E74C3C"
                elif severity == 'high' or severity == 'hoch':
                    color = "E67E22"
                else:
                    color = "F39C12"
                
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws_anomalies[f'{col}{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    ws_anomalies[f'{col}{row}'].font = Font(color="FFFFFF")
                
                row += 1
        
        # === Sheet 4: Rohdaten ===
        ws_raw = wb.create_sheet("Rohdaten")
        
        ws_raw['A1'] = "Ereignis-Log"
        ws_raw['A1'].font = Font(size=14, bold=True)
        
        headers = ["Zeit (ms)", "Pin", "Aktion", "Zyklus"]
        for col, header in enumerate(headers, start=1):
            cell = ws_raw.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        row = 4
        for entry in event_log[:1000]:  # Limit für Excel
            ws_raw[f'A{row}'] = f"{entry.get('time', 0):.2f}"
            ws_raw[f'B{row}'] = str(entry.get('pin', '-'))
            ws_raw[f'C{row}'] = str(entry.get('action', '-'))
            ws_raw[f'D{row}'] = str(entry.get('cycle', '-'))
            row += 1
        
        # Spaltenbreiten anpassen
        for ws in [ws_summary, ws_analysis, ws_anomalies, ws_raw]:
            for column in ws.columns:
                max_length = 0
                try:
                    column_letter = column[0].column_letter
                except AttributeError:
                    # MergedCell hat kein column_letter - überspringen
                    continue
                    
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Speichern
        wb.save(file_path)
    
    # ===== VERGLEICHSBERICHTE =====
    @staticmethod
    def generate_comparison_html(run_details_list: list, analysis_results: list) -> str:
        """
        Generiert HTML für Vergleichsbericht.
        """
        html = """
        <html>
        <head>
            <style>
                body { font-family: 'Segoe UI', sans-serif; background: #1e1e1e; color: #ecf0f1; margin: 20px; }
                h1 { color: #3498db; border-bottom: 3px solid #3498db; padding-bottom: 10px; }
                table { width: 100%; border-collapse: collapse; margin: 20px 0; }
                th { background: #34495e; color: #ecf0f1; padding: 12px; text-align: left; }
                td { background: #2c3e50; color: #ecf0f1; padding: 10px; border-bottom: 1px solid #34495e; }
            </style>
        </head>
        <body>
            <h1>Vergleichsbericht - Testläufe</h1>
            <table>
                <tr>
                    <th>Testlauf</th>
                    <th>Ø Zykluszeit (ms)</th>
                    <th>Stabilität (%)</th>
                    <th>Anomalien</th>
                    <th>Gesamt-Score</th>
                </tr>
        """
        
        for run_details, analysis in zip(run_details_list, analysis_results):
            cycle_stats = analysis.get('cycle_analysis', {})
            quality = analysis.get('quality_metrics', {})
            
            html += f"""
                <tr>
                    <td>#{run_details.get('id')} - {run_details.get('name', 'Unbekannt')}</td>
                    <td>{cycle_stats.get('avg', 0):.2f}</td>
                    <td>{cycle_stats.get('stability', 0):.1f}</td>
                    <td>{len(cycle_stats.get('anomalies', []))}</td>
                    <td>{quality.get('overall_score', 0):.1f}</td>
                </tr>
            """
        
        html += """
            </table>
        </body>
        </html>
        """
        
        return html
    
    @staticmethod
    def create_comparison_chart(run_details_list: list, analysis_results: list) -> io.BytesIO:
        """
        Erstellt Vergleichsdiagramm für mehrere Testläufe.
        """
        if not run_details_list or not analysis_results:
            return None
        
        plt.style.use('dark_background')
        
        labels = [f"#{d['id']}" for d in run_details_list]
        
        # Daten extrahieren
        avg_times = [a.get('cycle_analysis', {}).get('avg', 0) for a in analysis_results]
        stabilities = [a.get('cycle_analysis', {}).get('stability', 0) for a in analysis_results]
        anomalies_counts = [len(a.get('cycle_analysis', {}).get('anomalies', [])) for a in analysis_results]
        overall_scores = [a.get('quality_metrics', {}).get('overall_score', 0) for a in analysis_results]
        
        x = np.arange(len(labels))
        width = 0.2
        
        # HIER IST DIE VERVOLLSTÄNDIGUNG:
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.suptitle('Vergleich der Testläufe', fontsize=16, fontweight='bold', color='white')
        
        # 1. Durchschnittszeit
        ax1.bar(x, avg_times, width, color='#3498db', alpha=0.8)
        ax1.set_ylabel('Zeit (ms)', color='white')
        ax1.set_title('Ø Zykluszeit', color='white')
        ax1.set_xticks(x)
        ax1.set_xticklabels(labels, rotation=45)
        ax1.grid(True, alpha=0.3, axis='y')
        
        # 2. Stabilität
        ax2.bar(x, stabilities, width, color='#2ecc71', alpha=0.8)
        ax2.set_ylabel('Stabilität (%)', color='white')
        ax2.set_title('Stabilität', color='white')
        ax2.set_xticks(x)
        ax2.set_xticklabels(labels, rotation=45)
        ax2.set_ylim(0, 105)
        ax2.grid(True, alpha=0.3, axis='y')
        
        # 3. Anomalien
        ax3.bar(x, anomalies_counts, width, color='#e74c3c', alpha=0.8)
        ax3.set_ylabel('Anzahl', color='white')
        ax3.set_title('Anomalien', color='white')
        ax3.set_xticks(x)
        ax3.set_xticklabels(labels, rotation=45)
        ax3.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax3.grid(True, alpha=0.3, axis='y')
        
        # 4. Gesamt-Score
        colors_score = ['#2ecc71' if s >= 85 else '#f39c12' if s >= 70 else '#e74c3c' for s in overall_scores]
        ax4.bar(x, overall_scores, width, color=colors_score, alpha=0.8)
        ax4.set_ylabel('Score', color='white')
        ax4.set_title('Gesamt-Score', color='white')
        ax4.set_xticks(x)
        ax4.set_xticklabels(labels, rotation=45)
        ax4.set_ylim(0, 105)
        ax4.grid(True, alpha=0.3, axis='y')
        
        plt.tight_layout()
        
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=120, transparent=True)
        buf.seek(0)
        plt.close(fig)
        
        return buf