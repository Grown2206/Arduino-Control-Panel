import io
from datetime import datetime
import numpy as np
import base64

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.chart import LineChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MaxNLocator
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

from .trend_analyzer import TrendAnalyzer


class ReportGenerator:
    """Erstellt umfassende PDF-, Excel- und Vergleichs-Berichte mit Diagrammen."""

    @staticmethod
    def _create_sensor_chart(sensors_raw):
        if not MATPLOTLIB_AVAILABLE or not sensors_raw:
            return None

        plt.style.use('dark_background')
        fig, ax1 = plt.subplots(figsize=(8, 4))

        temp_times = [s['time_ms'] / 1000 for s in sensors_raw if s.get('sensor') == 'B24_TEMP' and 'time_ms' in s]
        temp_values = [s['value'] for s in sensors_raw if s.get('sensor') == 'B24_TEMP' and 'time_ms' in s]
        
        humid_times = [s['time_ms'] / 1000 for s in sensors_raw if s.get('sensor') == 'B24_HUMIDITY' and 'time_ms' in s]
        humid_values = [s['value'] for s in sensors_raw if s.get('sensor') == 'B24_HUMIDITY' and 'time_ms' in s]

        color_temp = '#e74c3c'
        ax1.set_xlabel('Zeit (s)', color='white')
        ax1.set_ylabel('Temperatur (°C)', color=color_temp)
        if temp_times:
            ax1.plot(temp_times, temp_values, color=color_temp, marker='.', linestyle='-', label='Temperatur')
        ax1.tick_params(axis='y', labelcolor=color_temp)
        ax1.tick_params(colors='white')

        ax2 = ax1.twinx()
        color_humid = '#3498db'
        ax2.set_ylabel('Luftfeuchtigkeit (%)', color=color_humid)
        if humid_times:
            ax2.plot(humid_times, humid_values, color=color_humid, marker='.', linestyle='-', label='Luftfeuchtigkeit')
        ax2.tick_params(axis='y', labelcolor=color_humid)

        ax1.set_title("Sensorverlauf während des Testlaufs", color='white')
        ax1.grid(True, linestyle='--', alpha=0.3)
        fig.tight_layout()
        
        lines, labels = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        if lines or lines2:
            ax2.legend(lines + lines2, labels + labels2, loc=0)

        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=150, transparent=True)
        buf.seek(0)
        plt.close(fig)
        return buf

    @staticmethod
    def _create_charts(analysis_result, sensors_raw=None):
        charts = []
        if not MATPLOTLIB_AVAILABLE or not analysis_result:
            return charts

        plt.style.use('dark_background')

        cycle_times = analysis_result.get('raw_cycle_times', [])
        if cycle_times:
            fig, ax = plt.subplots(figsize=(8, 4))
            ax.plot(cycle_times, marker='o', linestyle='-', color='#3498db')
            ax.set_title("Verlauf der Zykluszeiten", color='white')
            ax.set_xlabel("Zyklus", color='white'); ax.set_ylabel("Zeit (ms)", color='white')
            ax.grid(True, linestyle='--', alpha=0.3)
            buf = io.BytesIO(); fig.savefig(buf, format='png', dpi=150, transparent=True); buf.seek(0)
            charts.append(buf)
            plt.close(fig)

        step_times = analysis_result.get('raw_step_times', {})
        if step_times:
            fig, ax = plt.subplots(figsize=(8, 5))
            valid_step_times = {k: v for k, v in step_times.items() if v}
            if valid_step_times:
                labels = [key.replace(":", " -> ") for key in valid_step_times.keys()]
                ax.boxplot(valid_step_times.values())
                ax.set_xticklabels(labels, rotation=45, ha="right")
                ax.set_title("Verteilung der Einzelschritt-Zeiten", color='white')
                ax.set_ylabel("Zeit (ms)", color='white')
                ax.grid(True, linestyle='--', alpha=0.3)
                fig.tight_layout()
                buf = io.BytesIO(); fig.savefig(buf, format='png', dpi=150, transparent=True); buf.seek(0)
                charts.append(buf)
            plt.close(fig)
        
        if sensors_raw:
            sensor_chart_buf = ReportGenerator._create_sensor_chart(sensors_raw)
            if sensor_chart_buf:
                charts.append(sensor_chart_buf)
            
        return charts

    @staticmethod
    def generate_html(run_details, analysis_result):
        
        def get_trend_color(trend):
            if trend == 'increasing': return 'color: #e74c3c;'
            if trend == 'decreasing': return 'color: #3498db;'
            return 'color: #2ecc71;'

        ca = analysis_result.get('cycle_analysis', {})
        sa = analysis_result.get('step_analysis', {})
        sensor_stats = run_details.get('log', {}).get('sensors', {})

        html = f"""
        <html><head>
            <style>
                body {{ font-family: sans-serif; background-color: #2b2b2b; color: #ecf0f1; }}
                h1, h2, h3 {{ color: #3498db; border-bottom: 1px solid #555; padding-bottom: 5px; }}
                table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
                th, td {{ border: 1px solid #555; padding: 8px; text-align: left; }}
                th {{ background-color: #34495e; }}
                .trend {{ {get_trend_color(ca.get('trend', 'N/A'))} font-weight: bold; }}
                .anomalies {{ color: #e74c3c; }}
            </style>
        </head><body>
            <h1>Prüfbericht: {run_details.get('name', '')}</h1>
            <h2>Übersicht</h2>
            <table>
                <tr><th>Testlauf ID</th><td>{run_details.get('id', '-')}</td></tr>
                <tr><th>Sequenz</th><td>{run_details.get('sequence_name', '-')}</td></tr>
                <tr><th>Startzeit</th><td>{run_details.get('start_time', '-')}</td></tr>
                <tr><th>Endzeit</th><td>{run_details.get('end_time', '-')}</td></tr>
                <tr><th>Dauer</th><td>{run_details.get('duration', 0):.2f} s</td></tr>
                <tr><th>Zyklen</th><td>{run_details.get('cycles', 0)}</td></tr>
                <tr><th>Status</th><td><b>{run_details.get('status', '-')}</b></td></tr>
            </table>"""

        if sensor_stats:
            html += "<h2>Sensor-Auswertung</h2><table><tr><th>Sensor</th><th>Min</th><th>Max</th><th>Durchschnitt</th></tr>"
            if 'temp' in sensor_stats:
                s = sensor_stats.get('temp', {})
                html += f"<tr><td>Temperatur</td><td>{s.get('min', 0):.1f}°C</td><td>{s.get('max', 0):.1f}°C</td><td>{s.get('avg', 0):.1f}°C</td></tr>"
            if 'humid' in sensor_stats:
                s = sensor_stats.get('humid', {})
                html += f"<tr><td>Luftfeuchtigkeit</td><td>{s.get('min', 0):.1f}%</td><td>{s.get('max', 0):.1f}%</td><td>{s.get('avg', 0):.1f}%</td></tr>"
            html += "</table>"
        
        html += f"""
            <h2>Analyse der Zykluszeiten</h2>
            <table>
                <tr><th>Ø Zykluszeit</th><td>{ca.get('avg', 0):.2f} ms</td></tr>
                <tr><th>Stabilität</th><td>{ca.get('stability', 0):.1f}%</td></tr>
                <tr><th>Trend</th><td class='trend'>{ca.get('trend', 'N/A')}</td></tr>
                <tr><th>Anomalien</th><td class='anomalies'>{len(ca.get('anomalies', []))}</td></tr>
            </table>"""
        
        if sa:
            html += "<h2>Detail-Analyse der Einzelschritte</h2><table><tr><th>Schritt</th><th>Avg (ms)</th><th>Std (ms)</th><th>Min/Max (ms)</th></tr>"
            for step_name, data in sa.items():
                html += f"""
                <tr>
                    <td>{step_name.replace(':', ' -> ')}</td>
                    <td>{data.get('avg', 0):.2f}</td>
                    <td>{data.get('std', 0):.2f}</td>
                    <td>{data.get('min', 0):.2f} / {data.get('max', 0):.2f}</td>
                </tr>"""
            html += "</table>"

        html += "</body></html>"
        return html

    @staticmethod
    def generate_comparison_html(run_details_list, analysis_results):
        html = """
        <html><head><style>
            body { font-family: sans-serif; background-color: #2b2b2b; color: #ecf0f1; }
            h1, h2 { color: #3498db; border-bottom: 1px solid #555; padding-bottom: 5px; }
            table { width: 100%; border-collapse: collapse; margin-bottom: 20px; font-size: 11px; }
            th, td { border: 1px solid #555; padding: 6px; text-align: left; }
            th { background-color: #34495e; }
        </style></head><body>
            <h1>Vergleichsbericht</h1>
            <h2>Kennzahlen im Überblick</h2>
            <table>
                <tr>
                    <th>Metrik</th>
        """
        for details in run_details_list:
            html += f"<th>Testlauf #{details['id']}: {details['name']}</th>"
        html += "</tr>"

        metrics = [
            ("Dauer (s)", lambda d, a: f"{d.get('duration', 0):.2f}"),
            ("Zyklen", lambda d, a: d.get('cycles', 0)),
            ("Status", lambda d, a: d.get('status', '-')),
            ("Ø Zykluszeit (ms)", lambda d, a: f"{a.get('cycle_analysis', {}).get('avg', 0):.2f}"),
            ("Min Zykluszeit (ms)", lambda d, a: f"{a.get('cycle_analysis', {}).get('min', 0):.2f}"),
            ("Max Zykluszeit (ms)", lambda d, a: f"{a.get('cycle_analysis', {}).get('max', 0):.2f}"),
            ("Standardabweichung (ms)", lambda d, a: f"{a.get('cycle_analysis', {}).get('std', 0):.2f}"),
            ("Stabilität (%)", lambda d, a: f"{a.get('cycle_analysis', {}).get('stability', 0):.1f}"),
            ("Anomalien", lambda d, a: len(a.get('cycle_analysis', {}).get('anomalies', []))),
        ]

        for metric_name, getter in metrics:
            html += f"<tr><td><b>{metric_name}</b></td>"
            for i, details in enumerate(run_details_list):
                html += f"<td>{getter(details, analysis_results[i])}</td>"
            html += "</tr>"
        html += "</table>"
        
        has_sensor_data = any(d.get('log', {}).get('sensors') for d in run_details_list)
        if has_sensor_data:
            html += "<h2>Sensor-Kennzahlen im Vergleich</h2><table>"
            html += "<tr><th>Metrik</th>"
            for details in run_details_list:
                html += f"<th>Testlauf #{details['id']}</th>"
            html += "</tr>"

            sensor_metrics = [
                ("Ø Temperatur (°C)", lambda d: f"{d.get('log', {}).get('sensors', {}).get('temp', {}).get('avg', ''):.1f}" if d.get('log', {}).get('sensors', {}).get('temp') else 'N/A'),
                ("Ø Luftfeuchtigkeit (%)", lambda d: f"{d.get('log', {}).get('sensors', {}).get('humid', {}).get('avg', ''):.1f}" if d.get('log', {}).get('sensors', {}).get('humid') else 'N/A'),
            ]
            
            for metric_name, getter in sensor_metrics:
                html += f"<tr><td><b>{metric_name}</b></td>"
                for details in run_details_list:
                    html += f"<td>{getter(details)}</td>"
                html += "</tr>"
            html += "</table>"

        html += "</body></html>"
        return html

    @staticmethod
    def create_comparison_chart(run_details_list, analysis_results):
        if not MATPLOTLIB_AVAILABLE or not run_details_list:
            return None
        
        plt.style.use('dark_background')
        
        labels = [f"#{d['id']}" for d in run_details_list]
        avg_times = [a.get('cycle_analysis', {}).get('avg', 0) for a in analysis_results]
        stabilities = [a.get('cycle_analysis', {}).get('stability', 0) for a in analysis_results]
        anomalies_counts = [len(a.get('cycle_analysis', {}).get('anomalies', [])) for a in analysis_results]

        x = np.arange(len(labels))
        width = 0.35

        fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(10, 12))
        fig.tight_layout(pad=5.0)

        # Balkendiagramm für Durchschnittszeit
        ax1.bar(x, avg_times, width, label='Ø Zykluszeit', color='#3498db')
        ax1.set_ylabel('Zeit (ms)', color='white')
        ax1.set_title('Vergleich der Ø Zykluszeiten', color='white')
        ax1.set_xticks(x)
        ax1.set_xticklabels(labels, rotation=0, ha="center")
        ax1.tick_params(colors='white')
        ax1.grid(True, linestyle='--', alpha=0.3, axis='y')

        # Balkendiagramm für Stabilität
        ax2.bar(x, stabilities, width, label='Stabilität', color='#2ecc71')
        ax2.set_ylabel('Stabilität (%)', color='white')
        ax2.set_title('Vergleich der Stabilität', color='white')
        ax2.set_xticks(x)
        ax2.set_xticklabels(labels, rotation=0, ha="center")
        ax2.set_ylim(0, 105)
        ax2.tick_params(colors='white')
        ax2.grid(True, linestyle='--', alpha=0.3, axis='y')
        
        # Balkendiagramm für Anomalien
        ax3.bar(x, anomalies_counts, width, label='Anomalien', color='#e74c3c')
        ax3.set_ylabel('Anzahl Anomalien', color='white')
        ax3.set_title('Vergleich der Anzahl an Anomalien', color='white')
        ax3.set_xticks(x)
        ax3.set_xticklabels(labels, rotation=0, ha="center")
        ax3.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax3.set_ylim(bottom=0)
        ax3.tick_params(colors='white')
        ax3.grid(True, linestyle='--', alpha=0.3, axis='y')
        
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=120, transparent=True)
        buf.seek(0)
        plt.close(fig)
        return buf

    @staticmethod
    def generate_pdf(run_details, file_path):
        if not REPORTLAB_AVAILABLE:
            raise Exception("ReportLab ist nicht installiert.")

        doc = SimpleDocTemplate(file_path, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
        styles = getSampleStyleSheet()
        story = []
        
        log_data = run_details.get('log', {})
        event_log = log_data.get('events', [])
        sensor_stats = log_data.get('sensors', {})
        analysis = TrendAnalyzer.analyze_timing(event_log)
        
        story.append(Paragraph(f"<b>Prüfbericht: {run_details.get('name', '')}</b>", styles['h1']))
        story.append(Spacer(1, 0.5 * cm))

        story.append(Paragraph("<b>Übersicht</b>", styles['h2']))
        summary_data = [
            ["Testlauf ID:", str(run_details.get('id', '-'))],
            ["Sequenz:", str(run_details.get('sequence_name', '-'))],
            ["Startzeit:", str(run_details.get('start_time', '-'))],
            ["Endzeit:", str(run_details.get('end_time', '-'))],
            ["Dauer:", f"{run_details.get('duration', 0):.2f} s"],
            ["Zyklen:", str(run_details.get('cycles', 0))],
            ["Status:", Paragraph(f"<b>{run_details.get('status', '-')}</b>", styles['Normal'])],
        ]
        story.append(Table(summary_data, colWidths=[4*cm, None], style=[('VALIGN', (0,0), (-1,-1), 'TOP')]))
        story.append(Spacer(1, 1 * cm))

        if sensor_stats:
            story.append(Paragraph("<b>Sensor-Auswertung</b>", styles['h2']))
            sensor_data = [["Sensor", "Min", "Max", "Durchschnitt"]]
            if 'temp' in sensor_stats:
                s = sensor_stats.get('temp', {})
                sensor_data.append(["Temperatur", f"{s.get('min', 0):.1f}°C", f"{s.get('max', 0):.1f}°C", f"{s.get('avg', 0):.1f}°C"])
            if 'humid' in sensor_stats:
                s = sensor_stats.get('humid', {})
                sensor_data.append(["Luftfeuchtigkeit", f"{s.get('min', 0):.1f}%", f"{s.get('max', 0):.1f}%", f"{s.get('avg', 0):.1f}%"])
            story.append(Table(sensor_data, style=[('GRID', (0,0), (-1,-1), 1, colors.black)]))
            story.append(Spacer(1, 1 * cm))

        ca = analysis.get('cycle_analysis', {})
        story.append(Paragraph("<b>Analyse der Zykluszeiten</b>", styles['h2']))
        cycle_data = [
            ["Durchschnitt:", f"{ca.get('avg', 0):.2f} ms"],
            ["Stabilität:", f"{ca.get('stability', 0):.1f}%"],
            ["Anomalien:", str(len(ca.get('anomalies', [])))], 
        ]
        story.append(Table(cycle_data))
        story.append(Spacer(1, 0.5 * cm))

        charts = ReportGenerator._create_charts(analysis, log_data.get('sensors_raw'))
        for chart_buf in charts:
            story.append(Image(chart_buf, width=16*cm, height='auto'))
            story.append(Spacer(1, 0.5 * cm))
            
        story.append(PageBreak())
        
        step_analysis = analysis.get('step_analysis', {})
        if step_analysis:
            story.append(Paragraph("<b>Detail-Analyse der Einzelschritte</b>", styles['h2']))
            for step_name, data in step_analysis.items():
                story.append(Paragraph(f"<b>{step_name.replace(':', ' -> ')}</b>", styles['h3']))
                table_data = [
                    ["Avg:", f"{data.get('avg', 0):.2f} ms"],
                    ["Std:", f"{data.get('std', 0):.2f} ms"],
                    ["Min/Max:", f"{data.get('min', 0):.2f} / {data.get('max', 0):.2f} ms"],
                ]
                story.append(Table(table_data))
                anomalies_list = data.get('anomalies', [])
                if anomalies_list:
                    story.append(Paragraph(f"<font color='red'>Anomalien gefunden: {str(len(anomalies_list))}</font>", styles['Normal']))
                story.append(Spacer(1, 0.5 * cm))

        story.append(Paragraph("<b>Ereignis-Log</b>", styles['h2']))
        if event_log:
            log_table_data = [["Zeit (ms)", "Pin", "Aktion", "Zyklus", "Schrittzeit (ms)"]]
            raw_step_deltas = analysis.get('raw_step_deltas', [])

            for i, entry in enumerate(event_log[:500]):
                step_delta = raw_step_deltas[i] if i < len(raw_step_deltas) else ''
                log_table_data.append([
                    f"{entry.get('time', 0):.2f}",
                    str(entry.get('pin', '-')),
                    str(entry.get('action', '-')),
                    str(entry.get('cycle', '-')),
                    f"{step_delta:.2f}" if isinstance(step_delta, (int, float)) else ''
                ])

            log_table = Table(log_table_data, colWidths=[2.5*cm, 2*cm, 3*cm, 2*cm, 3*cm])
            log_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.grey), ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 8), ('GRID', (0,0), (-1,-1), 0.5, colors.black)
            ]))
            story.append(log_table)

        doc.build(story)

    @staticmethod
    def generate_excel(run_details, file_path):
        if not EXCEL_AVAILABLE:
            raise Exception("openpyxl ist nicht installiert.")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Zusammenfassung"
        
        log_data = run_details.get('log', {})
        event_log = log_data.get('events', [])
        sensor_stats = log_data.get('sensors', {})
        analysis = TrendAnalyzer.analyze_timing(event_log)

        ws['A1'] = f"Prüfbericht: {run_details.get('name', '')}"; ws['A1'].font = Font(size=14, bold=True)
        
        summary = [
            ("Testlauf ID", run_details.get('id')), 
            ("Sequenz", run_details.get('sequence_name')),
            ("Startzeit", run_details.get('start_time')), 
            ("Endzeit", run_details.get('end_time')),
            ("Dauer (s)", f"{run_details.get('duration', 0):.2f}"),
            ("Zyklen", run_details.get('cycles', 0)),
            ("Status", run_details.get('status')),
        ]
        for i, (key, value) in enumerate(summary, 3):
            ws[f'A{i}'] = key; ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = str(value) if value is not None else ""

        current_row = ws.max_row + 2
        
        if sensor_stats:
            ws[f'A{current_row}'] = "Sensor-Auswertung"; ws[f'A{current_row}'].font = Font(size=12, bold=True)
            current_row += 1
            ws.append(["Sensor", "Min", "Max", "Durchschnitt"])
            if 'temp' in sensor_stats:
                s = sensor_stats.get('temp', {})
                ws.append(["Temperatur", f"{s.get('min', 0):.1f}°C", f"{s.get('max', 0):.1f}°C", f"{s.get('avg', 0):.1f}°C"])
            if 'humid' in sensor_stats:
                s = sensor_stats.get('humid', {})
                ws.append(["Luftfeuchtigkeit", f"{s.get('min', 0):.1f}%", f"{s.get('max', 0):.1f}%", f"{s.get('avg', 0):.1f}%"])
            current_row = ws.max_row + 2


        ws[f'D3'] = "Analyse der Zykluszeiten"; ws[f'D3'].font = Font(size=12, bold=True)
        ca = analysis.get('cycle_analysis', {})
        analysis_summary = [
            ("Ø Zykluszeit (ms)", f"{ca.get('avg', 0):.2f}"),
            ("Stabilität", f"{ca.get('stability', 0):.1f}%"),
            ("Trend", ca.get('trend', 'N/A')),
            ("Anomalien (Zyklen)", len(ca.get('anomalies', []))),
        ]
        for i, (key, value) in enumerate(analysis_summary, 4):
            ws[f'D{i}'] = key; ws[f'D{i}'].font = Font(bold=True)
            ws[f'E{i}'] = value

        ws_details = wb.create_sheet("Details & Rohdaten")

        ws_details['A1'] = "Detail-Analyse der Einzelschritte"; ws_details['A1'].font = Font(size=12, bold=True)
        step_analysis = analysis.get('step_analysis', {})
        if step_analysis:
            ws_details.append(["Schritt", "Avg (ms)", "Std (ms)", "Min/Max (ms)"])
            for step_name, data in step_analysis.items():
                ws_details.append([
                    step_name.replace(':', ' -> '),
                    f"{data.get('avg', 0):.2f}",
                    f"{data.get('std', 0):.2f}",
                    f"{data.get('min', 0):.2f} / {data.get('max', 0):.2f}"
                ])
        
        ws_details.append([])
        
        log_start_row = ws_details.max_row + 1
        ws_details[f'A{log_start_row}'] = "Ereignis-Log & Zykluszeiten"; ws_details[f'A{log_start_row}'].font = Font(size=12, bold=True)
        
        headers = ["Zeit (ms)", "Pin", "Aktion", "Zyklus", "Schrittzeit (ms)", "", "Zyklus-Nr.", "Zykluszeit (ms)"]
        ws_details.append(headers)
        
        raw_step_deltas = analysis.get('raw_step_deltas', [])
        raw_cycle_times = analysis.get('raw_cycle_times', [])

        for i, entry in enumerate(event_log):
            row_data = [
                entry.get('time'), entry.get('pin'), entry.get('action'),
                entry.get('cycle'), raw_step_deltas[i] if i < len(raw_step_deltas) else ''
            ]
            if i < len(raw_cycle_times):
                row_data += ["", i + 1, raw_cycle_times[i]]
            
            ws_details.append(row_data)
        
        if raw_cycle_times:
            chart = LineChart()
            chart.title = "Verlauf der Zykluszeiten"
            chart.x_axis.title = "Zyklus"; chart.y_axis.title = "Zeit (ms)"
            chart_start_row = log_start_row + 2
            max_row_chart = len(raw_cycle_times) + chart_start_row -1
            
            data = Reference(ws_details, min_col=8, min_row=chart_start_row, max_row=max_row_chart)
            cats = Reference(ws_details, min_col=7, min_row=chart_start_row, max_row=max_row_chart)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            ws.add_chart(chart, "J2")
            
        for col_letter in ['A', 'B', 'D', 'E']: ws.column_dimensions[col_letter].width = 25
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'G', 'H']: ws_details.column_dimensions[col_letter].width = 20
        
        wb.save(file_path)

